# AUTHORSHIP NOTICE - DO NOT EDIT OR REMOVE
# Author: Tanishq Bhelonde - Intern Finance - Fortis Healthcare(HHPL)
# Proprietary for Fortis Healthcare(HHPL)
# Date: 04/04/25
# Run the following in Terminal - pip install pandas openpyxl

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

def process_excel(excel_file_path, output_dir="."):
    """Process Excel file and generate doctor statements"""
    if not os.path.exists(excel_file_path):
        print(f"Error: File {excel_file_path} not found.")
        return
    
    try:
        print(f"Reading data from {excel_file_path}...")
        df = pd.read_excel(excel_file_path)
        df.columns = df.columns.str.strip().str.replace(' ', '_').str.lower()
        
        
        column_mapping = {
            'category': 'category', 'ipid': 'ipid', 'patient_name': 'patient_name',
            'bill_date': 'bill_date', 'surgery': 'surgery', 'doctor_name': 'doctor_name',
            'net_amount': 'net_amount', 'doctor_share': 'doctor_share', 
            'agreed_share': 'agreed_share', 'gross_payable': 'gross_payable',
            'tds_10%': 'tds_10_percent', 'net_payable': 'net_payable',
            'payor_name': 'payor_name'
        }
        
        
        columns_to_use = {col: internal_col for standard_col, internal_col in column_mapping.items() 
                          for col in df.columns if standard_col.lower() in col.lower()}
        
        if len(columns_to_use) < len(column_mapping):
            print("Warning: Not all expected columns found in Excel file.")
            print("Available columns:", df.columns.tolist())
            print("Missing columns:", set(column_mapping.values()) - set(columns_to_use.values()))
        
        
        df.rename(columns=columns_to_use, inplace=True, errors='ignore')
        
        
        text_cols = ['patient_name', 'doctor_name', 'category', 'surgery', 'payor_name']
        for col in text_cols:
            if col in df.columns and df[col].dtype == 'object':
                df[col] = df[col].str.replace('_', ' ').str.title()
        
        
        for col in [c for c in df.columns if 'date' in c.lower()]:
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            except:
                print(f"Warning: Could not convert {col} to datetime format.")
        
        
        if 'doctor_name' not in df.columns or 'bill_date' not in df.columns:
            print("Error: Required columns missing (doctor_name or bill_date)")
            return
        
        
        df_copy = df.copy()
        try:
            df_copy['month_year'] = df_copy['bill_date'].dt.strftime('%Y-%m')
        except:
            print("Error: Unable to process bill_date")
            return
        
        doctors = df_copy['doctor_name'].unique()
        print(f"Creating statement files for {len(doctors)} doctors...")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        files_created = []
        
        
        header_green = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        
        categories = ["IP Procedure", "IP Bedside Procedure", "IP Assist", "IP Investigation",
                      "IP Physio", "IP Visit", "OP Consultation", "OP Bedside Procedure",
                      "OP Procedures", "OP Physio", "EHC"]
        
        
        for doctor in doctors:
            doctor_data = df_copy[df_copy['doctor_name'] == doctor].copy()
            months = sorted(doctor_data['month_year'].unique())
            
            safe_name = str(doctor).replace('.', '').replace(' ', '_').lower()
            filename = os.path.join(output_dir, f"doctor_statement_{safe_name}_{timestamp}.xlsx")
            
            try:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    for month in months:
                        month_data = doctor_data[doctor_data['month_year'] == month].copy()
                        
                        
                        display_month = month.split('-')[1]
                        display_year = month.split('-')[0][2:]
                        formatted_month = f"{display_month}-{display_year}"
                        
                        
                        category_data = {cat: [] for cat in categories}
                        
                        if 'category' in month_data.columns:
                            for _, row in month_data.iterrows():
                                cat = str(row.get('category', '')).strip().lower()
                                
                                
                                if 'ip' in cat and ('procedure' in cat or 'surgery' in cat):
                                    category = 'IP Procedure'
                                elif 'ip' in cat and 'beside' in cat:
                                    category = 'IP Bedside Procedure'
                                elif 'ip' in cat and 'assist' in cat:
                                    category = 'IP Assist'
                                elif 'ip' in cat and ('investigation' in cat or 'test' in cat):
                                    category = 'IP Investigation'
                                elif 'ip' in cat and 'physio' in cat:
                                    category = 'IP Physio'
                                elif 'ip' in cat and 'visit' in cat:
                                    category = 'IP Visit'
                                elif ('op' in cat or 'opd' in cat) and ('consult' in cat):
                                    category = 'OP Consultation'
                                elif ('op' in cat or 'opd' in cat) and 'beside' in cat:
                                    category = 'OP Bedside Procedure'
                                elif ('op' in cat or 'opd' in cat) and ('procedure' in cat or 'surgery' in cat):
                                    category = 'OP Procedures'
                                elif ('op' in cat or 'opd' in cat) and 'physio' in cat:
                                    category = 'OP Physio'
                                elif 'ehc' in cat or 'ech' in cat:
                                    category = 'EHC'
                                else:
                                    category = 'IP Bedside Procedure' if 'ip' in cat else 'OP Bedside Procedure' if 'op' in cat or 'opd' in cat else 'EHC'
                                
                                if category in category_data:
                                    category_data[category].append(row)
                        
                        
                        statement_rows = []
                        for category in categories:
                            rows = category_data.get(category, [])
                            if rows:
                                df_cat = pd.DataFrame(rows)
                                gross = df_cat['gross_payable'].sum() if 'gross_payable' in df_cat.columns else 0
                                tds = df_cat['tds_10_percent'].sum() if 'tds_10_percent' in df_cat.columns else 0
                                net = df_cat['net_payable'].sum() if 'net_payable' in df_cat.columns else 0
                            else:
                                gross, tds, net = 0, 0, 0
                                
                            statement_rows.append({
                                'Particular': category,
                                'Gross Payable': gross,
                                'TDS 10%': tds,
                                'Net Payable': net
                            })
                        
                        
                        statement_rows.append({
                            'Particular': 'Total',
                            'Gross Payable': sum(row['Gross Payable'] for row in statement_rows),
                            'TDS 10%': sum(row['TDS 10%'] for row in statement_rows),
                            'Net Payable': sum(row['Net Payable'] for row in statement_rows)
                        })
                        
                        
                        sheet_name = month
                        if sheet_name not in writer.book.sheetnames:
                            writer.book.create_sheet(sheet_name)
                        ws = writer.book[sheet_name]
                        
                        
                        ws.merge_cells('A1:D1')
                        ws['A1'] = "Hiranandani Healthcare Pvt. Ltd. (Fortis Vashi)"
                        ws['A1'].font = Font(bold=True, color="FFFFFF")
                        ws['A1'].fill = header_green
                        ws['A1'].alignment = Alignment(horizontal='center')
                        
                        ws.merge_cells('A2:D2')
                        ws['A2'] = f"Statement for the Month of {formatted_month}"
                        ws['A2'].alignment = Alignment(horizontal='center')
                        
                        
                        for cell_ref, value in [('A3', 'Particular'), ('B3', 'Gross Payable'), 
                                               ('C3', 'TDS 10%'), ('D3', 'Net Payable')]:
                            ws[cell_ref] = value
                            ws[cell_ref].font = Font(bold=True, color="FFFFFF")
                            ws[cell_ref].fill = header_green
                            ws[cell_ref].alignment = Alignment(horizontal='center')
                            ws[cell_ref].border = thin_border
                        
                        
                        ws.column_dimensions['A'].width = 25
                        for col in ['B', 'C', 'D']:
                            ws.column_dimensions[col].width = 15
                        
                        
                        for i, row in enumerate(statement_rows):
                            row_num = i + 4
                            is_total_row = row['Particular'] == 'Total'
                            
                            ws[f'A{row_num}'] = row['Particular']
                            ws[f'B{row_num}'] = row['Gross Payable'] if row['Gross Payable'] != 0 else '-'
                            ws[f'C{row_num}'] = row['TDS 10%'] if row['TDS 10%'] != 0 else '-'
                            ws[f'D{row_num}'] = row['Net Payable'] if row['Net Payable'] != 0 else '-'
                            
                            for col in ['A', 'B', 'C', 'D']:
                                ws[f'{col}{row_num}'].border = thin_border
                                if is_total_row:
                                    ws[f'{col}{row_num}'].fill = header_green
                                    ws[f'{col}{row_num}'].font = Font(bold=True, color="FFFFFF")
                        
                        
                        detail_start_row = len(statement_rows) + 6
                        ws.merge_cells(f'A{detail_start_row}:M{detail_start_row}')
                        ws[f'A{detail_start_row}'] = "Detailed Report"
                        ws[f'A{detail_start_row}'].font = Font(bold=True)
                        
                        
                        detail_cols = ['Category', 'IPID', 'Patient Name', 'Bill Date', 
                                      'Surgery/Procedure/Package', 'Doctor Name', 'Net Amount', 
                                      'Doctor Share', 'Agreed Share', 'Gross Payable', 
                                      'TDS 10%', 'Net Payable', 'Payor Name']
                        
                        
                        header_row = detail_start_row + 1
                        for col_idx, col_name in enumerate(detail_cols):
                            col_letter = get_column_letter(col_idx + 1)
                            ws[f'{col_letter}{header_row}'] = col_name
                            ws[f'{col_letter}{header_row}'].font = Font(bold=True, color="FFFFFF")
                            ws[f'{col_letter}{header_row}'].fill = header_green
                            ws[f'{col_letter}{header_row}'].border = thin_border
                            ws[f'{col_letter}{header_row}'].alignment = Alignment(horizontal='center')
                            ws.column_dimensions[col_letter].width = 15
                        
                        
                        if not month_data.empty:
                            col_map = {
                                'category': 'Category', 'ipid': 'IPID', 
                                'patient_name': 'Patient Name', 'bill_date': 'Bill Date',
                                'surgery': 'Surgery/Procedure/Package', 'doctor_name': 'Doctor Name',
                                'net_amount': 'Net Amount', 'doctor_share': 'Doctor Share',
                                'agreed_share': 'Agreed Share', 'gross_payable': 'Gross Payable',
                                'tds_10_percent': 'TDS 10%', 'net_payable': 'Net Payable',
                                'payor_name': 'Payor Name'
                            }
                            
                            sorted_data = month_data.sort_values('bill_date')
                            for i, (_, data_row) in enumerate(sorted_data.iterrows()):
                                row_num = header_row + i + 1
                                
                                for col_idx, col_name in enumerate(detail_cols):
                                    col_letter = get_column_letter(col_idx + 1)
                                    df_col = next((k for k, v in col_map.items() 
                                                  if v == col_name and k in data_row.index), None)
                                    
                                    if df_col is not None:
                                        value = data_row[df_col]
                                        
                                        if df_col == 'bill_date' and pd.notna(value) and isinstance(value, pd.Timestamp):
                                            value = value.strftime('%Y-%m-%d')
                                        
                                        if df_col in ['gross_payable', 'tds_10_percent', 'net_payable',
                                                    'net_amount', 'doctor_share', 'agreed_share'] and pd.notna(value) and value == 0:
                                            value = '-'
                                        
                                        ws[f'{col_letter}{row_num}'] = value
                                    else:
                                        ws[f'{col_letter}{row_num}'] = ""
                                    
                                    ws[f'{col_letter}{row_num}'].border = thin_border
                
                files_created.append(filename)
                print(f"  - Created statement for {doctor} with {len(months)} month(s)")
                
            except Exception as e:
                print(f"  - Error creating statement for {doctor}: {str(e)}")
        
        print(f"Generated {len(files_created)} doctor statement files")
        return files_created
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

if __name__ == "__main__":
    print("Excel Data Extraction Tool for Medical Billing")
    print("---------------------------------------------")
    
    excel_file_path = input("Enter the path to the consolidated Excel file: ")
    output_dir = input("Enter the output directory path (press Enter to use current directory): ") or "."
    
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"Created output directory: {output_dir}")
        except Exception as e:
            print(f"Error creating directory: {str(e)}")
            output_dir = "."
    
    process_excel(excel_file_path, output_dir) 